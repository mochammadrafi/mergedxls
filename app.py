import os
from openpyxl import load_workbook

source_files = os.listdir('files');

target_file = "output.xlsx"
target_workbook = load_workbook(target_file)
target_worksheet = target_workbook.worksheets[0]

for file in source_files:
    load_file = load_workbook('files/' + file)
    load_worksheet = load_file.worksheets[0]
    print("MERGED FILE " + file)

    max_row = load_worksheet.max_row
    max_column = load_worksheet.max_column
    for i in range (1, max_row + 1):
        value_row = []
        if i > 1:
            for j in range (1, max_column + 1):
                data = load_worksheet.cell(row = i, column = j).value
                value_row.append(data)
        if value_row:
            target_worksheet.append(value_row)

target_workbook.save(str(target_file))